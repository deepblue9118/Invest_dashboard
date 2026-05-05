#!/usr/bin/env python3
"""GitHub Actions용 대쉬보드 생성기 - Google Drive Excel -> index.html"""
import os, subprocess, sys
from datetime import datetime

FILE_ID   = "1uNSAGSsZlD8_Tl2c7yoW0FEDVFbwhN9W"
XLSX      = "/tmp/invest.xlsx"
OUT       = "index.html"
today_str = datetime.now().strftime("%Y%m%d")

# 패키지 설치
subprocess.run([sys.executable, "-m", "pip", "install", "gdown", "openpyxl", "-q"], check=True)
import gdown
from openpyxl import load_workbook

# Excel 다운로드
print("Google Drive에서 Excel 다운로드...")
gdown.download("https://drive.google.com/uc?id=" + FILE_ID, XLSX, quiet=False)

# 데이터 추출
def v(x):  return x if x is not None else 0
def vp(x): return x if x is not None else 0.0

wb   = load_workbook(XLSX, data_only=True)
ws_h = wb["현황"]; ws_t = wb["토스증권"]; ws_r = wb["퇴직연금"]
ws_p = wb["개인연금"]; ws_k = wb["김찬호"]; ws_d = wb["배당금"]

d = {}
d["total_asset"]  = v(ws_h["B4"].value)
d["toss_cur"]     = v(ws_h["B5"].value)
d["retire_cur"]   = v(ws_h["B6"].value)
d["pension_cur"]  = v(ws_h["B7"].value)
d["growth_amt"]   = v(ws_h["E5"].value)
d["div_amt"]      = v(ws_h["E6"].value)
d["stock_amt"]    = v(ws_h["E7"].value)
d["stable_amt"]   = v(ws_h["E8"].value)
d["toss_cost"]    = v(ws_t["E4"].value)
d["toss_gain"]    = v(ws_t["K4"].value)
d["toss_ret"]     = vp(ws_t["L4"].value)
d["toss_tr"]      = v(ws_t["N4"].value)
d["toss_tr_ret"]  = vp(ws_t["O4"].value)
d["toss_div_26"]  = v(ws_t["M4"].value)
d["retire_cost"]  = v(ws_r["H5"].value)
d["retire_gain"]  = v(ws_r["L5"].value)
d["retire_ret"]   = vp(ws_r["M5"].value)
d["retire_risk"]  = vp(ws_r["P2"].value)
d["pension_gain"] = v(ws_p["J5"].value)
d["pension_cost"] = d["pension_cur"] - d["pension_gain"]
d["pension_ret"]  = (d["pension_cur"] / d["pension_cost"] - 1) if d["pension_cost"] else 0
d["kc_cur"]       = v(ws_k["K4"].value)
d["kc_cost"]      = v(ws_k["H4"].value)
d["kc_gain"]      = v(ws_k["L4"].value)
d["kc_ret"]       = vp(ws_k["M4"].value)
d["total_cost"]   = d["toss_cost"] + d["retire_cost"] + d["pension_cost"]
d["total_gain"]   = d["toss_gain"] + d["retire_gain"] + d["pension_gain"]
d["total_ret"]    = d["total_gain"] / d["total_cost"] if d["total_cost"] else 0
d["div_26"]       = v(ws_d["D3"].value)
d["div_25"]       = v(ws_d["E3"].value)
d["div_monthly"]  = round(d["div_26"] / 4) if d["div_26"] else 0
d["div_annual"]   = d["div_26"] * 3

toss_rows = []
for r in range(5, 12):
    rk = ws_t.cell(r, 1).value
    nm = ws_t.cell(r, 2).value
    if rk and nm:
        toss_rows.append({
            "rank": int(rk), "name": nm,
            "qty":  v(ws_t.cell(r, 3).value),
            "cost": v(ws_t.cell(r, 5).value),
            "cur":  v(ws_t.cell(r, 9).value),
            "gain": v(ws_t.cell(r, 11).value),
            "ret":  vp(ws_t.cell(r, 12).value),
            "div":  v(ws_t.cell(r, 13).value),
        })

retire_rows = []
for r in range(6, 14):
    no = ws_r.cell(r, 1).value
    nm = ws_r.cell(r, 2).value
    if no and nm:
        retire_rows.append({
            "no": int(no), "name": nm,
            "qty": v(ws_r.cell(r, 6).value), "buy": v(ws_r.cell(r, 7).value),
            "cur_price": v(ws_r.cell(r, 9).value),
            "cost": v(ws_r.cell(r, 8).value), "cur": v(ws_r.cell(r, 11).value),
            "gain": v(ws_r.cell(r, 12).value), "ret": vp(ws_r.cell(r, 13).value),
            "tgt":  vp(ws_r.cell(r, 4).value), "now": vp(ws_r.cell(r, 14).value),
            "diff": vp(ws_r.cell(r, 15).value), "adj": v(ws_r.cell(r, 16).value),
        })

pension_rows = []
for r in range(6, 10):
    no = ws_p.cell(r, 1).value
    nm = ws_p.cell(r, 2).value
    if no and nm:
        qty = v(ws_p.cell(r, 5).value)
        buy = v(ws_p.cell(r, 6).value)
        pension_rows.append({
            "no": int(no), "name": nm, "qty": qty, "buy": buy,
            "cur_price": v(ws_p.cell(r, 7).value), "cost": qty * buy,
            "cur": v(ws_p.cell(r, 9).value), "gain": v(ws_p.cell(r, 10).value),
            "ret": vp(ws_p.cell(r, 11).value), "tgt": vp(ws_p.cell(r, 4).value),
            "now": vp(ws_p.cell(r, 12).value), "diff": vp(ws_p.cell(r, 13).value),
            "adj": v(ws_p.cell(r, 14).value),
        })

kc_rows = []
for r in range(5, 10):
    no = ws_k.cell(r, 1).value
    nm = ws_k.cell(r, 2).value
    if no and nm:
        kc_rows.append({
            "no": int(no), "name": nm,
            "qty": v(ws_k.cell(r, 6).value), "buy": v(ws_k.cell(r, 7).value),
            "cur_price": v(ws_k.cell(r, 9).value),
            "cost": v(ws_k.cell(r, 8).value), "cur": v(ws_k.cell(r, 11).value),
            "gain": v(ws_k.cell(r, 12).value), "ret": vp(ws_k.cell(r, 13).value),
            "tgt":  vp(ws_k.cell(r, 4).value), "now": vp(ws_k.cell(r, 14).value),
            "diff": vp(ws_k.cell(r, 15).value), "adj": v(ws_k.cell(r, 16).value),
        })

div_rows = []
for r in range(4, 20):
    rk = ws_d.cell(r, 1).value
    nm = ws_d.cell(r, 2).value
    if rk and nm:
        div_rows.append({
            "rank": int(rk), "name": nm,
            "acct": ws_d.cell(r, 3).value or "",
            "d26":  v(ws_d.cell(r, 4).value),
            "d25":  v(ws_d.cell(r, 5).value),
            "note": ws_d.cell(r, 6).value or "",
        })

goals = []
for row in toss_rows:
    nm = row["name"]
    price = row["cur"] / row["qty"] if row["qty"] else 0
    if nm == "JEPQ":
        goals.append({"name": "JEPQ", "cur": row["qty"], "target": 1000, "daily": 120000, "price": price})
    elif nm == "SPYI":
        goals.append({"name": "SPYI", "cur": row["qty"], "target": 400,  "daily": 50000,  "price": price})
    elif nm == "SCHD":
        goals.append({"name": "SCHD", "cur": row["qty"], "target": 2000, "daily": 0,      "price": price})

# 유틸리티
def fmt(n, sign=False):
    if not n: return "—"
    try:
        n = float(n)
        return ("+{:,.0f}".format(n) if n >= 0 else "{:,.0f}".format(n)) if sign else "{:,.0f}".format(n)
    except:
        return "—"

def fmtp(n, sign=False):
    if n is None: return "—"
    try:
        n = float(n)
        return ("+{:.2%}".format(n) if n >= 0 else "{:.2%}".format(n)) if sign else "{:.2%}".format(n)
    except:
        return "—"

def clr(n, pos="var(--green)", neg="var(--red)", zero="var(--text)"):
    try:
        n = float(n or 0)
        return pos if n > 0 else (neg if n < 0 else zero)
    except:
        return zero

def bar(ratio, length=20):
    r = max(0.0, min(1.0, ratio))
    f = int(r * length)
    return "█" * f + "░" * (length - f)

# HTML 조각 생성 (f-string 안에 백슬래시/복잡한 표현 없이)
goal_html = ""
for g in goals:
    cur    = g["cur"]
    tgt    = g["target"]
    daily  = g["daily"]
    price  = g["price"]
    ratio  = cur / tgt if tgt else 0
    remain = max(0, tgt - cur)
    if daily > 0 and price > 0:
        eta = "약 {}개월 후".format(max(1, round(remain * price / daily / 30)))
    else:
        eta = "배당재투자"
    daily_str  = "배당재투자" if not daily else "{:,.0f}".format(daily)
    ratio_clr  = clr(ratio - 0.5, "var(--green)", "var(--orange)", "var(--orange)")
    ratio_str  = fmtp(ratio)
    bar_str    = bar(ratio)
    cur_str    = "{:,.2f}".format(cur)
    tgt_str    = "{:,.0f}".format(tgt)
    remain_str = "{:,.2f}".format(remain)
    goal_html += (
        "<tr>"
        "<td class='name'>" + g["name"] + "</td>"
        "<td>" + cur_str + "</td>"
        "<td>" + tgt_str + "</td>"
        "<td style='color:" + ratio_clr + "'><b>" + ratio_str + "</b></td>"
        "<td class='bar-cell'><span style='color:var(--blue);font-family:monospace'>" + bar_str + "</span></td>"
        "<td>" + remain_str + "</td>"
        "<td style='color:var(--blue)'>" + eta + "</td>"
        "<td>" + daily_str + "</td>"
        "</tr>"
    )

def hold_rows(rows):
    html = ""
    for i, r in enumerate(rows):
        bg       = "#f8faff" if i % 2 == 0 else "#fff"
        gain_clr = clr(r["gain"])
        ret_clr  = clr(r["ret"])
        html += (
            "<tr style='background:" + bg + "'>"
            "<td style='text-align:center'>" + str(r["no"]) + "</td>"
            "<td class='name'>" + r["name"] + "</td>"
            "<td>" + "{:,.0f}".format(r["qty"]) + "</td>"
            "<td>" + fmt(r["buy"]) + "</td>"
            "<td>" + fmt(r["cur_price"]) + "</td>"
            "<td>" + fmt(r["cost"]) + "</td>"
            "<td><b>" + fmt(r["cur"]) + "</b></td>"
            "<td style='color:" + gain_clr + "'><b>" + fmt(r["gain"], True) + "</b></td>"
            "<td style='color:" + ret_clr + "'><b>" + fmtp(r["ret"], True) + "</b></td>"
            "</tr>"
        )
    return html

def rebal_rows(rows, extra_col=False):
    html = ""
    for i, r in enumerate(rows):
        diff   = r["diff"]
        bg     = "#fff0f0" if abs(diff) > 0.05 else ("#f8faff" if i % 2 == 0 else "#fff")
        dc     = "var(--red)" if abs(diff) > 0.05 else clr(-diff, "var(--green)", "var(--orange)")
        ac     = "var(--red)" if r["adj"] < 0 else "var(--green)"
        flag   = "🔴" if abs(diff) > 0.05 else ("🟢" if abs(diff) <= 0.02 else "🟡")
        action = "매도" if diff > 0 else "매수"
        tgt_s  = fmtp(r["tgt"]) if r["tgt"] else "—"
        now_s  = fmtp(r["now"])
        diff_s = fmtp(diff, True)
        adj_s  = fmt(r["adj"], True) if r["adj"] else "—"
        html += (
            "<tr style='background:" + bg + "'>"
            "<td style='text-align:center'>" + str(r["no"]) + "</td>"
            "<td class='name'>" + r["name"] + "</td>"
            "<td>" + tgt_s + "</td>"
            "<td style='color:var(--blue)'><b>" + now_s + "</b></td>"
            "<td style='color:" + dc + "'><b>" + diff_s + "</b></td>"
            "<td style='color:" + ac + "'>" + adj_s + "</td>"
            "<td>" + action + "</td>"
        )
        if extra_col:
            html += "<td>" + flag + "</td>"
        html += "</tr>"
    return html

def div_rows_html(rows):
    html = ""
    for i, r in enumerate(rows):
        bg    = "#f8faff" if i % 2 == 0 else "#fff"
        yoy   = r["d26"] - r["d25"] if r["d26"] and r["d25"] else None
        yoy_c = clr(yoy) if yoy is not None else "var(--text)"
        yoy_s = fmt(yoy, True) if yoy is not None else "—"
        d26_s = fmt(r["d26"]) if r["d26"] else "—"
        d25_s = fmt(r["d25"]) if r["d25"] else "—"
        html += (
            "<tr style='background:" + bg + "'>"
            "<td style='text-align:center'>" + str(r["rank"]) + "</td>"
            "<td class='name'>" + r["name"] + "</td>"
            "<td>" + r["acct"] + "</td>"
            "<td style='color:var(--blue)'><b>" + d26_s + "</b></td>"
            "<td>" + d25_s + "</td>"
            "<td style='color:" + yoy_c + "'>" + yoy_s + "</td>"
            "<td style='font-size:0.8em;color:#666'>" + r["note"] + "</td>"
            "</tr>"
        )
    return html

total_asset = d["total_asset"]

acct_rows_html = ""
for name, cost, cur, gain, ret in [
    ("토스증권", d["toss_cost"],   d["toss_cur"],   d["toss_gain"],   d["toss_ret"]),
    ("퇴직연금", d["retire_cost"], d["retire_cur"], d["retire_gain"], d["retire_ret"]),
    ("개인연금", d["pension_cost"],d["pension_cur"],d["pension_gain"],d["pension_ret"]),
]:
    wt      = cur / total_asset if total_asset else 0
    gc      = clr(gain)
    rc      = clr(ret)
    acct_rows_html += (
        "<tr>"
        "<td class='name'><b>" + name + "</b></td>"
        "<td>" + fmt(cost) + "</td>"
        "<td><b>" + fmt(cur) + "</b></td>"
        "<td style='color:" + gc + "'><b>" + fmt(gain, True) + "</b></td>"
        "<td style='color:" + rc + "'><b>" + fmtp(ret, True) + "</b></td>"
        "<td>{:.1%}</td>".format(wt) +
        "</tr>"
    )

alloc_html = ""
for name, amt, color in [
    ("성장형",       d["growth_amt"],  "#3b82f6"),
    ("배당/커버드콜", d["div_amt"],     "#22c55e"),
    ("안정자산",     d["stable_amt"],   "#f97316"),
    ("개별주식",     d["stock_amt"],    "#a78bfa"),
]:
    ratio = amt / total_asset if total_asset else 0
    w     = max(2, int(ratio * 100))
    alloc_html += (
        "<div style='display:flex;align-items:center;gap:8px;margin-bottom:6px'>"
        "<div style='width:90px;font-size:0.85em'>" + name + "</div>"
        "<div style='flex:1;background:#e5e7eb;border-radius:4px;height:20px;overflow:hidden'>"
        "<div style='width:" + str(w) + "%;background:" + color + ";height:100%;border-radius:4px;display:flex;align-items:center;padding-left:6px'>"
        "<span style='color:white;font-size:0.75em;font-weight:bold'>{:.1%}</span>".format(ratio) +
        "</div></div>"
        "<div style='width:100px;text-align:right;font-size:0.85em'>" + fmt(amt) + "</div>"
        "</div>"
    )

risk      = d["retire_risk"]
risk_diff = risk - 0.70
risk_c    = "#dc2626" if abs(risk_diff) > 0.05 else "#16a34a"

ta_clr   = clr(d["total_gain"])
tr_clr   = clr(d["total_ret"])
kc_clr   = clr(d["kc_ret"])
ttr_clr  = clr(d["toss_tr_ret"])
rag_clr  = clr(d["retire_gain"])
rar_clr  = clr(d["retire_ret"])
pag_clr  = clr(d["pension_gain"])
par_clr  = clr(d["pension_ret"])
kag_clr  = clr(d["kc_gain"])
kar_clr  = clr(d["kc_ret"])
dyoy_clr = clr(d["div_26"] - d["div_25"])

HTML = """<!DOCTYPE html>
<html lang="ko">
<head>
<meta charset="UTF-8">
<meta name="viewport" content="width=device-width,initial-scale=1.0">
<title>주식투자 대쉬보드 ({today})</title>
<style>
:root{{--blue:#2563eb;--green:#16a34a;--red:#dc2626;--orange:#ea580c;--text:#1e293b;--bg:#f1f5f9;--card:#fff;--border:#e2e8f0;--navy:#1e3a5f}}
*{{box-sizing:border-box;margin:0;padding:0}}
body{{font-family:'Malgun Gothic','맑은 고딕',sans-serif;background:var(--bg);color:var(--text);font-size:13px}}
.container{{max-width:1400px;margin:0 auto;padding:16px}}
.header{{background:var(--navy);color:white;padding:16px 20px;border-radius:12px;margin-bottom:16px}}
.header h1{{font-size:1.4em;margin-bottom:4px}}.header p{{font-size:0.85em;opacity:0.8}}
.grid2{{display:grid;grid-template-columns:1fr 1fr;gap:16px;margin-bottom:16px}}
.grid4{{display:grid;grid-template-columns:repeat(4,1fr);gap:12px;margin-bottom:16px}}
.card{{background:var(--card);border-radius:10px;padding:16px;border:1px solid var(--border);box-shadow:0 1px 3px rgba(0,0,0,.06)}}
.section{{background:var(--card);border-radius:10px;border:1px solid var(--border);margin-bottom:16px;overflow:hidden}}
.sh{{background:var(--navy);color:white;padding:10px 16px;font-weight:bold;font-size:.95em}}
.sb{{padding:16px}}
.kl{{font-size:.78em;color:#64748b;margin-bottom:4px}}
.kv{{font-size:1.5em;font-weight:bold;color:var(--navy)}}
.ks{{font-size:.85em;margin-top:2px}}
table{{width:100%;border-collapse:collapse;font-size:.88em}}
th{{background:#1e3a5f;color:white;padding:7px 8px;text-align:right;font-weight:600}}
th:first-child,th.name{{text-align:left}}
td{{padding:6px 8px;border-bottom:1px solid var(--border);text-align:right}}
td:first-child,td.name{{text-align:left}}
tr:last-child td{{border-bottom:none}}
.tr td{{background:#fef3c7!important;font-weight:bold;border-top:2px solid #d97706}}
.bar-cell{{font-family:monospace;font-size:.85em}}
@media(max-width:768px){{.grid2,.grid4{{grid-template-columns:1fr}}}}
</style>
</head>
<body><div class="container">

<div class="header">
  <h1>&#128202; 주식투자 포트폴리오 대쉬보드</h1>
  <p>업데이트: {today4}.{today46}.{today6} &nbsp;|&nbsp; 원본: Google Drive 251130_주식투자 현황.xlsx &nbsp;|&nbsp; 매일 자동 갱신</p>
</div>

<div class="grid4">
  <div class="card"><div class="kl">총 평가금액 (본인 3계좌)</div><div class="kv">{total_asset}</div><div class="ks" style="color:#64748b">원금: {total_cost}</div></div>
  <div class="card"><div class="kl">총 손익 / 수익률</div><div class="kv" style="color:{ta_clr}">{total_gain}</div><div class="ks" style="color:{tr_clr}">{total_ret}</div></div>
  <div class="card"><div class="kl">26년 배당금 누적</div><div class="kv" style="color:var(--blue)">{div_26}</div><div class="ks" style="color:#64748b">월 평균: {div_monthly}</div></div>
  <div class="card"><div class="kl">25년 연간 배당금</div><div class="kv" style="color:var(--blue)">{div_25}</div><div class="ks" style="color:#64748b">연환산 추정(x3): {div_annual}</div></div>
</div>

<div class="grid2" style="margin-bottom:16px">
  <div class="card" style="background:#fffbeb;border-color:#fde68a">
    <div class="kl">&#128102; 김찬호 계좌 (별도 관리, 합산 제외)</div>
    <div style="display:flex;gap:20px;margin-top:4px">
      <div><span style="color:#64748b;font-size:.8em">평가금액</span><br><b>{kc_cur}</b></div>
      <div><span style="color:#64748b;font-size:.8em">투자원금</span><br><b>{kc_cost}</b></div>
      <div><span style="color:#64748b;font-size:.8em">수익률</span><br><b style="color:{kc_clr}">{kc_ret}</b></div>
    </div>
  </div>
  <div class="card" style="background:#eff6ff;border-color:#bfdbfe">
    <div class="kl">&#128200; 토스증권 TR 수익률 (배당금 포함)</div>
    <div style="display:flex;gap:20px;margin-top:4px">
      <div><span style="color:#64748b;font-size:.8em">TR 평가금액</span><br><b>{toss_tr}</b></div>
      <div><span style="color:#64748b;font-size:.8em">TR 수익률</span><br><b style="color:{ttr_clr}">{toss_tr_ret}</b></div>
    </div>
  </div>
</div>

<div class="grid2">
  <div class="section"><div class="sh">&#9313; 계좌별 현황 요약</div><div class="sb" style="padding:0"><table>
    <tr><th class="name">계좌</th><th>투자원금</th><th>현재금액</th><th>손익</th><th>수익률</th><th>비중</th></tr>
    {acct_rows}
    <tr class="tr"><td class="name">합계</td><td>{total_cost}</td><td><b>{total_asset}</b></td>
      <td style="color:{ta_clr}"><b>{total_gain}</b></td>
      <td style="color:{tr_clr}"><b>{total_ret}</b></td><td>100.0%</td></tr>
  </table></div></div>
  <div class="section"><div class="sh">&#9313; 자산 유형 배분</div><div class="sb">
    {alloc}
    <div style="margin-top:12px;padding:10px;background:#f0f9ff;border-radius:8px;border:1px solid #bae6fd">
      <span style="font-size:.85em">퇴직연금 위험자산 비중: </span>
      <b style="color:{risk_c}">{risk}</b>
      <span style="font-size:.82em;color:#64748b"> (목표 70% / 차이 {risk_diff})</span>
    </div>
  </div></div>
</div>

<div class="section"><div class="sh">&#9314; 목표 달성 현황 (JEPQ · SPYI · SCHD)</div><div class="sb" style="padding:0"><table>
  <tr><th class="name">종목</th><th>현재주수</th><th>목표주수</th><th>달성률</th><th>진행바</th><th>남은주수</th><th>예상달성</th><th>일일투자</th></tr>
  {goals}
</table></div></div>

<div class="section"><div class="sh">&#9315; 배당금 현황 (종목별)</div><div class="sb" style="padding:0"><table>
  <tr><th style="text-align:center">순위</th><th class="name">종목명</th><th class="name">계좌</th><th>26년</th><th>25년</th><th>YoY</th><th class="name">비고</th></tr>
  {divs}
  <tr class="tr"><td></td><td class="name">합계</td><td></td>
    <td style="color:var(--blue)"><b>{div_26}</b></td>
    <td><b>{div_25}</b></td>
    <td style="color:{dyoy_clr}"><b>{div_yoy}</b></td><td></td></tr>
</table></div></div>

<div class="section"><div class="sh">&#9316; 퇴직연금 보유 현황</div><div class="sb" style="padding:0"><table>
  <tr><th style="text-align:center">순번</th><th class="name">종목명</th><th>수량</th><th>매입단가</th><th>현재단가</th><th>매입금액</th><th>현재금액</th><th>손익</th><th>수익률</th></tr>
  {retire_hold}
  <tr class="tr"><td></td><td class="name">합계</td><td></td><td></td><td></td>
    <td>{retire_cost}</td><td><b>{retire_cur}</b></td>
    <td style="color:{rag_clr}"><b>{retire_gain}</b></td>
    <td style="color:{rar_clr}"><b>{retire_ret}</b></td></tr>
</table></div></div>

<div class="section"><div class="sh">&#9317; 리밸런싱 현황 (&#128308; = &#177;5% 초과)</div><div class="sb">
  <p style="font-weight:bold;margin-bottom:8px;color:var(--navy)">[ 퇴직연금 ]</p>
  <table style="margin-bottom:16px">
    <tr><th style="text-align:center">순번</th><th class="name">종목명</th><th>목표</th><th>현재</th><th>차이</th><th>조정금액</th><th>매수/매도</th><th>상태</th></tr>
    {rebal_r}
  </table>
  <p style="font-weight:bold;margin-bottom:8px;color:var(--navy)">[ 개인연금 ]</p>
  <table style="margin-bottom:16px">
    <tr><th style="text-align:center">순번</th><th class="name">종목명</th><th>목표</th><th>현재</th><th>차이</th><th>조정금액</th><th>매수/매도</th></tr>
    {rebal_p}
  </table>
  <p style="font-weight:bold;margin-bottom:8px;color:var(--navy)">[ 김찬호 계좌 ]</p>
  <table>
    <tr><th style="text-align:center">순번</th><th class="name">종목명</th><th>목표</th><th>현재</th><th>차이</th><th>조정금액</th><th>매수/매도</th></tr>
    {rebal_k}
  </table>
</div></div>

<div class="section"><div class="sh">&#9318; 개인연금 보유 현황</div><div class="sb" style="padding:0"><table>
  <tr><th style="text-align:center">순번</th><th class="name">종목명</th><th>수량</th><th>매입단가</th><th>현재단가</th><th>매입금액</th><th>현재금액</th><th>손익</th><th>수익률</th></tr>
  {pension_hold}
  <tr class="tr"><td></td><td class="name">합계</td><td></td><td></td><td></td>
    <td>{pension_cost}</td><td><b>{pension_cur}</b></td>
    <td style="color:{pag_clr}"><b>{pension_gain}</b></td>
    <td style="color:{par_clr}"><b>{pension_ret}</b></td></tr>
</table></div></div>

<div class="section"><div class="sh">&#9319; 김찬호 계좌 (아들 계좌, 별도 관리)</div><div class="sb" style="padding:0"><table>
  <tr><th style="text-align:center">순번</th><th class="name">종목명</th><th>수량</th><th>매입단가</th><th>현재단가</th><th>매입금액</th><th>현재금액</th><th>손익</th><th>수익률</th></tr>
  {kc_hold}
  <tr class="tr"><td></td><td class="name">합계</td><td></td><td></td><td></td>
    <td>{kc_cost}</td><td><b>{kc_cur}</b></td>
    <td style="color:{kag_clr}"><b>{kc_gain}</b></td>
    <td style="color:{kar_clr}"><b>{kc_ret2}</b></td></tr>
</table>
<p style="font-size:.8em;color:#64748b;margin-top:10px">※ KODEX 200(20%) -&gt; 장기적으로 나스닥100 또는 전세계 인덱스(ACWI)로 교체 검토 권장</p>
</div></div>

<div style="text-align:center;color:#94a3b8;font-size:.8em;padding:16px 0">
  자동 생성: {today4}-{today46}-{today6} | Google Drive -&gt; GitHub Actions -&gt; GitHub Pages
</div>
</div></body></html>""".format(
    today=today_str,
    today4=today_str[:4], today46=today_str[4:6], today6=today_str[6:],
    total_asset  = fmt(d["total_asset"]),
    total_cost   = fmt(d["total_cost"]),
    total_gain   = fmt(d["total_gain"], True),
    total_ret    = fmtp(d["total_ret"], True),
    ta_clr=ta_clr, tr_clr=tr_clr,
    div_26       = fmt(d["div_26"]),
    div_25       = fmt(d["div_25"]),
    div_monthly  = fmt(d["div_monthly"]),
    div_annual   = fmt(d["div_annual"]),
    div_yoy      = fmt(d["div_26"] - d["div_25"], True),
    dyoy_clr     = dyoy_clr,
    kc_cur       = fmt(d["kc_cur"]),
    kc_cost      = fmt(d["kc_cost"]),
    kc_gain      = fmt(d["kc_gain"], True),
    kc_ret       = fmtp(d["kc_ret"], True),
    kc_clr       = kc_clr,
    toss_tr      = fmt(d["toss_tr"]),
    toss_tr_ret  = fmtp(d["toss_tr_ret"], True),
    ttr_clr      = ttr_clr,
    acct_rows    = acct_rows_html,
    alloc        = alloc_html,
    risk         = fmtp(risk),
    risk_diff    = fmtp(risk_diff, True),
    risk_c       = risk_c,
    goals        = goal_html,
    divs         = div_rows_html(div_rows),
    retire_hold  = hold_rows(retire_rows),
    retire_cost  = fmt(d["retire_cost"]),
    retire_cur   = fmt(d["retire_cur"]),
    retire_gain  = fmt(d["retire_gain"], True),
    retire_ret   = fmtp(d["retire_ret"], True),
    rag_clr=rag_clr, rar_clr=rar_clr,
    rebal_r      = rebal_rows(retire_rows, extra_col=True),
    rebal_p      = rebal_rows(pension_rows),
    rebal_k      = rebal_rows(kc_rows),
    pension_hold = hold_rows(pension_rows),
    pension_cost = fmt(d["pension_cost"]),
    pension_cur  = fmt(d["pension_cur"]),
    pension_gain = fmt(d["pension_gain"], True),
    pension_ret  = fmtp(d["pension_ret"], True),
    pag_clr=pag_clr, par_clr=par_clr,
    kc_hold      = hold_rows(kc_rows),
    kc_ret2      = fmtp(d["kc_ret"], True),
    kag_clr=kag_clr, kar_clr=kar_clr,
)

with open(OUT, "w", encoding="utf-8") as f:
    f.write(HTML)
print("index.html 생성 완료 ({})".format(today_str))
print("총 자산: {}원 | 수익률: {} | 배당: {}원".format(
    fmt(d["total_asset"]), fmtp(d["total_ret"], True), fmt(d["div_26"])))
